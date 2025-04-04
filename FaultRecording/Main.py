from PyQt5 import QtWidgets, QtCore
from MainWindow import Ui_MainWindow
import speech_recognition as sr
import threading
from openpyxl import Workbook, load_workbook
import os
import time

class FaultRecorderApp(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setup_connections()
        self.recognizer = sr.Recognizer()
        self.current_machine_name = ""
        self.current_fault_description = ""
        self.remaining_time = 0
        self.timer = None

    def setup_connections(self):
        self.ui.comboBox.currentIndexChanged.connect(self.dummy_method)
        self.ui.speakButton.clicked.connect(self.start_speech_recognition)
        self.ui.buttonBox.accepted.connect(self.handle_yes_response)
        self.ui.buttonBox.rejected.connect(self.handle_no_response)
        self.ui.saveButton.clicked.connect(self.save_manual_entry)
        self.ui.exitButton.clicked.connect(self.close_application)
        self.ui.manualEntryButton.clicked.connect(self.show_manual_entry_ui)

    def dummy_method(self):
        pass

    def start_speech_recognition(self):
        self.ui.resultOutput.setText("Mikrofon kontrolleri yapılıyor...")
        if not self.is_microphone_available():
            self.ui.resultOutput.setText("Mikrofon bulunamadı. Lütfen bir mikrofon bağlayın.")
            return

        self.toggle_ui_elements(False)
        self.remaining_time = 10  # Makine adı için 10 saniye
        self.timer = QtCore.QTimer()
        self.timer.timeout.connect(self.update_timer)
        self.timer.start(1000)  # Her saniye güncelle

        threading.Thread(target=self.capture_machine_name, daemon=True).start()

    def is_microphone_available(self):
        try:
            mic_list = sr.Microphone.list_microphone_names()
            return len(mic_list) > 0
        except OSError:
            return False

    def update_timer(self):
        if self.remaining_time > 0:
            self.remaining_time -= 1
            if self.ui.checkSpeechOutput.text() == "Makine adı doğru mu?":
                QtCore.QMetaObject.invokeMethod(self.ui.resultOutput, "setText", 
                                                QtCore.Qt.QueuedConnection, 
                                                QtCore.Q_ARG(str, f"Lütfen makine adını söyleyin... Kalan süre: {self.remaining_time} saniye"))
            elif self.ui.checkSpeechOutput.text() == "Arıza tanımı doğru mu?":
                QtCore.QMetaObject.invokeMethod(self.ui.resultOutput, "setText", 
                                                QtCore.Qt.QueuedConnection, 
                                                QtCore.Q_ARG(str, f"Lütfen arıza tanımını söyleyin... Kalan süre: {self.remaining_time} saniye"))
        else:
            self.timer.stop()

    def capture_machine_name(self):
        try:
            start_time = time.time()
            with sr.Microphone() as source:
                self.recognizer.adjust_for_ambient_noise(source, duration=2)
                QtCore.QMetaObject.invokeMethod(self.ui.resultOutput, "setText", 
                                                QtCore.Qt.QueuedConnection, 
                                                QtCore.Q_ARG(str, "Lütfen makine adını söyleyin..."))
                audio = self.recognizer.listen(source, timeout=5, phrase_time_limit=10)
                machine_name = self.recognizer.recognize_google(audio, language="tr-TR")
                end_time = time.time()
                duration = end_time - start_time

                with open("transcription_usage.txt", "a", encoding="utf-8") as f:
                    f.write(f"{QtCore.QDateTime.currentDateTime().toString('yyyy-MM-dd HH:mm:ss')} - Machine Name - {duration:.2f} saniye\n")

                self.check_transcription_usage()

                self.current_machine_name = machine_name
                QtCore.QMetaObject.invokeMethod(self.ui.speechOutput, "setText", 
                                                QtCore.Qt.QueuedConnection, 
                                                QtCore.Q_ARG(str, f"Makine adı: {machine_name}"))
                QtCore.QMetaObject.invokeMethod(self.ui.checkSpeechOutput, "setText", 
                                                QtCore.Qt.QueuedConnection, 
                                                QtCore.Q_ARG(str, "Makine adı doğru mu?"))
                QtCore.QMetaObject.invokeMethod(self.ui.checkSpeechOutput, "show", 
                                                QtCore.Qt.QueuedConnection)
                QtCore.QMetaObject.invokeMethod(self.ui.buttonBox, "show", 
                                                QtCore.Qt.QueuedConnection)
        except sr.UnknownValueError:
            QtCore.QMetaObject.invokeMethod(self.ui.resultOutput, "setText", 
                                            QtCore.Qt.QueuedConnection, 
                                            QtCore.Q_ARG(str, "Makine adı anlaşılamadı. Tekrar denemek için 'Konuş' butonuna basın veya manuel giriş yapın."))
        except sr.RequestError:
            QtCore.QMetaObject.invokeMethod(self.ui.resultOutput, "setText", 
                                            QtCore.Qt.QueuedConnection, 
                                            QtCore.Q_ARG(str, "Bağlantı veya API hatası. İnternet bağlantınızı kontrol edin."))
        except sr.WaitTimeoutError:
            QtCore.QMetaObject.invokeMethod(self.ui.resultOutput, "setText", 
                                            QtCore.Qt.QueuedConnection, 
                                            QtCore.Q_ARG(str, "Belirtilen süre boyunca konuşma olmadı. Tekrar denemek için 'Konuş' butonuna basın."))

    def handle_yes_response(self):
        if self.ui.checkSpeechOutput.text() == "Makine adı doğru mu?":
            self.ui.checkSpeechOutput.hide()
            self.ui.buttonBox.hide()
            self.ui.resultOutput.setText("Mikrofon kontrolleri yapılıyor...")
            self.current_fault_description = ""
            self.remaining_time = 30  # Arıza tanımı için 30 saniye
            self.timer.start(1000)
            threading.Thread(target=self.capture_fault_description, daemon=True).start()
        elif self.ui.checkSpeechOutput.text() == "Arıza tanımı doğru mu?":
            self.save_to_excel(self.current_machine_name, self.current_fault_description)
            self.reset_ui()

    def handle_no_response(self):
        if self.ui.checkSpeechOutput.text() == "Makine adı doğru mu?":
            self.ui.resultOutput.setText("Lütfen makine adını tekrar söyleyin.")
            self.current_machine_name = ""
            self.ui.speechOutput.clear()
            self.ui.checkSpeechOutput.hide()
            self.ui.buttonBox.hide()
            self.remaining_time = 10
            self.timer.start(1000)
            threading.Thread(target=self.capture_machine_name, daemon=True).start()
        elif self.ui.checkSpeechOutput.text() == "Arıza tanımı doğru mu?":
            self.ui.resultOutput.setText("Lütfen arıza tanımını tekrar söyleyin.")
            self.current_fault_description = ""
            self.ui.speechOutput.clear()
            self.ui.checkSpeechOutput.hide()
            self.ui.buttonBox.hide()
            self.remaining_time = 30
            self.timer.start(1000)
            threading.Thread(target=self.capture_fault_description, daemon=True).start()

    def capture_fault_description(self):
        try:
            start_time = time.time()
            with sr.Microphone() as source:
                self.recognizer.adjust_for_ambient_noise(source, duration=2)
                QtCore.QMetaObject.invokeMethod(self.ui.resultOutput, "setText", 
                                                QtCore.Qt.QueuedConnection, 
                                                QtCore.Q_ARG(str, "Lütfen arıza tanımını söyleyin..."))
                audio = self.recognizer.listen(source, timeout=5, phrase_time_limit=30)
                fault_description = self.recognizer.recognize_google(audio, language="tr-TR")
                end_time = time.time()
                duration = end_time - start_time

                with open("transcription_usage.txt", "a", encoding="utf-8") as f:
                    f.write(f"{QtCore.QDateTime.currentDateTime().toString('yyyy-MM-dd HH:mm:ss')} - Fault Description - {duration:.2f} saniye\n")

                self.check_transcription_usage()

                self.current_fault_description = fault_description
                QtCore.QMetaObject.invokeMethod(self.ui.speechOutput, "setText", 
                                                QtCore.Qt.QueuedConnection, 
                                                QtCore.Q_ARG(str, f"Arıza tanımı: {fault_description}"))
                QtCore.QMetaObject.invokeMethod(self.ui.checkSpeechOutput, "setText", 
                                                QtCore.Qt.QueuedConnection, 
                                                QtCore.Q_ARG(str, "Arıza tanımı doğru mu?"))
                QtCore.QMetaObject.invokeMethod(self.ui.checkSpeechOutput, "show", 
                                                QtCore.Qt.QueuedConnection)
                QtCore.QMetaObject.invokeMethod(self.ui.buttonBox, "show", 
                                                QtCore.Qt.QueuedConnection)
        except sr.UnknownValueError:
            QtCore.QMetaObject.invokeMethod(self.ui.resultOutput, "setText", 
                                            QtCore.Qt.QueuedConnection, 
                                            QtCore.Q_ARG(str, "Arıza tanımı anlaşılamadı. Tekrar denemek için 'Konuş' butonuna basın veya manuel giriş yapın."))
        except sr.RequestError:
            QtCore.QMetaObject.invokeMethod(self.ui.resultOutput, "setText", 
                                            QtCore.Qt.QueuedConnection, 
                                            QtCore.Q_ARG(str, "Bağlantı veya API hatası. İnternet bağlantınızı kontrol edin."))
        except sr.WaitTimeoutError:
            QtCore.QMetaObject.invokeMethod(self.ui.resultOutput, "setText", 
                                            QtCore.Qt.QueuedConnection, 
                                            QtCore.Q_ARG(str, "Belirtilen süre boyunca konuşma olmadı. Tekrar denemek için 'Konuş' butonuna basın."))

    def save_manual_entry(self):
        machine_name = self.ui.comboBox.currentText()
        fault_description = self.ui.faultDefinitionManualEntry.toPlainText().strip()
        if machine_name and fault_description:
            try:
                self.save_to_excel(machine_name, fault_description)
                self.ui.faultDefinitionManualEntry.clear()
            except Exception as e:
                self.ui.resultOutput.setText(f"Excel dosyasına yazma hatası: {e}")
        else:
            self.ui.resultOutput.setText("Lütfen tüm alanları doldurun.")

    def save_to_excel(self, machine_name, fault_description):
        directory = "C:/BakimKayitlari"
        file_path = f"{directory}/ariza_kayitlari_{QtCore.QDate.currentDate().toString('yyyyMMdd')}.xlsx"

        try:
            if not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True)
        except OSError as e:
            self.ui.resultOutput.setText(f"Dosya dizini oluşturulamadı: {e}")
            return

        try:
            if not os.path.exists(file_path):
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Fault Records"
                sheet.append(["Timestamp", "Machine Name", "Fault Description"])
            else:
                workbook = load_workbook(file_path)
                sheet = workbook.active

            timestamp = QtCore.QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm:ss")
            sheet.append([timestamp, machine_name, fault_description])
            workbook.save(file_path)
            self.ui.resultOutput.setText(f"Kayıt başarıyla eklendi: {machine_name} - {fault_description}")
        except PermissionError:
            self.ui.resultOutput.setText("Dosya yazma izni yok. Lütfen Excel dosyasını kapatın veya başka bir dosya adı kullanmayı deneyin.")
        except Exception as e:
            self.ui.resultOutput.setText(f"Excel dosyasına yazma hatası: {e}")

    def check_transcription_usage(self):
        total_duration = 0
        if os.path.exists("transcription_usage.txt"):
            with open("transcription_usage.txt", "r", encoding="utf-8") as f:
                for line in f:
                    if "saniye" in line:
                        duration = float(line.split(" - ")[-1].replace(" saniye", ""))
                        total_duration += duration
        if total_duration >= 3600:  # 60 dakika = 3600 saniye
            self.ui.resultOutput.setText("Uyarı: Aylık 60 dakikalık ücretsiz kota doldu! Daha fazla kullanım ücretlendirilebilir.")
        return total_duration

    def toggle_ui_elements(self, show_manual=True):
        self.ui.manualEntryText.setVisible(show_manual)
        self.ui.machineName.setVisible(show_manual)
        self.ui.faultDefinition.setVisible(show_manual)
        self.ui.comboBox.setVisible(show_manual)
        self.ui.faultDefinitionManualEntry.setVisible(show_manual)
        self.ui.saveButton.setVisible(show_manual)

    def reset_ui(self):
        self.ui.speechOutput.clear()
        self.ui.checkSpeechOutput.setText("")
        self.ui.checkSpeechOutput.hide()
        self.ui.buttonBox.hide()
        self.ui.resultOutput.setText("Kayıt başarıyla eklendi!")
        self.toggle_ui_elements(False)

    def show_manual_entry_ui(self):
        if self.ui.manualEntryText.isVisible():
            self.ui.resultOutput.setText("Manuel giriş ekranı kapatıldı.")
            self.toggle_ui_elements(False)
        else:
            self.ui.resultOutput.setText("Manuel giriş ekranı açıldı.")
            self.toggle_ui_elements(True)

    def close_application(self):
        self.close()

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    window = FaultRecorderApp()
    window.show()
    sys.exit(app.exec_())